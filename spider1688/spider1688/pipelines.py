# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface

# from spider1688.spiders.al_category import AlCategorySpider
# from ..spider1688.spiders.al_category import AlCategorySpider
from .spiders.al_category import AlCategorySpider
from itemadapter import ItemAdapter
import openpyxl
import pymysql

# mark -> 41
row = 2


class CategoryPipeline(object):
    def __init__(self):
        self.db = None
        self.cursor = None

    def open_spider(self, spider):
        if isinstance(spider, AlCategorySpider):
            # 建立数据库连接
            # self.db = pymysql.connect(host='localhost',
            #                           user='root',
            #                           passwd='12345',
            #                           port=3306,
            #                           db='test',
            #                           charset='utf8')
            # self.cursor = self.db.cursor()
            pass

    def process_item(self, item, spider):
        # insert for mysql
        # self.cursor.execute('insert into test1(title,key2,) values(("%s"),("%s")) % (title[0],key2[0])')
        # results = (self.cursor.fetchall())
        # print(results)

        # 如果数据条目不对则尝试回滚
        # len(item['merch_price']) != len(item['merch_num']) | len(item['merch_price']) != len(item['merch_name'])

        # 插入row行数据 start from 2nd
        # ->in
        global row

        wb = openpyxl.load_workbook(r"C:\Users\11246\Desktop\test.xlsx")
        sheet = wb['sheet1']
        # for column in sheet.max_column:        # insert max column datas
        sheet.cell(row=row, column=1, value=item['company_name'])
        sheet.cell(row=row, column=2, value=item['company_address'])
        sheet.cell(row=row, column=3, value=item['company_area'])
        sheet.cell(row=row, column=4, value=item['company_city'])
        sheet.cell(row=row, column=5, value=item['company_archive'])
        sheet.cell(row=row, column=6, value=item['company_service'])
        sheet.cell(row=row, column=7, value=item['company_device'])
        sheet.cell(row=row, column=8, value=item['company_turnover'])
        sheet.cell(row=row, column=9, value=item['company_staff'])
        sheet.cell(row=row, column=10, value=item['company_merchandise'])
        sheet.cell(row=row, column=11, value=item['company_detail_url'])
        sheet.cell(row=row, column=12, value=item['company_merch_url'])
        sheet.cell(row=row, column=13, value=item['company_detail_picture'])
        for i in range(len(item['merch_name'])):
            sheet.cell(row=row, column=14, value=str(item['merch_name'][i]))
            sheet.cell(row=row, column=15, value=str(item['merch_num'][i]))
            sheet.cell(row=row, column=16, value=str(item['merch_price'][i]))
            row += 1

        row += 1
        wb.save(r"C:\Users\11246\Desktop\test.xlsx")

        # sheet.cell(行, 列).value = 数据
        # 保存,传入原文件则在原文件上追加数据，也可以保存为新文

        return item

    def close_spider(self, spider):
        pass
        # 关闭数据库连接
        # if isinstance(spider, AlCategorySpider):
        #     # 游标对象断开连接
        #     self.cursor.close()
        #
        #     # 数据库对象断开连接
        #     self.db.close()


