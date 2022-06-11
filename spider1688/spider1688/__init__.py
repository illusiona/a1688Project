import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\11246\Desktop\test.xlsx")
sheet = wb['sheet1']
for row in range(10, 100):
    sheet.cell(row=row, column=1, value='asdsa')
    sheet.cell(row=row, column=2, value='asdsa')
    sheet.cell(row=row, column=3, value='asdsa')
    sheet.cell(row=row, column=4, value='asdsa')
    sheet.cell(row=row, column=5, value='asdsa')
    sheet.cell(row=row, column=6, value='asdsa')
    sheet.cell(row=row, column=7, value='asdsa')
    sheet.cell(row=row, column=8, value='asdsa')
    sheet.cell(row=row, column=9, value='asdsa')
    sheet.cell(row=row, column=10, value='asdsa')
    sheet.cell(row=row, column=11, value='asdsa')
    sheet.cell(row=row, column=12, value='asdsa')
    sheet.cell(row=row, column=13, value='asdsa')
    # sheet.cell(row=row, column=14, value=str(item['merch_name']))
    # sheet.cell(row=row, column=15, value=str(item['merch_num']))
    # sheet.cell(row=row, column=16, value=str(item['merch_price']))

    row += 1
    wb.save(r"C:\Users\11246\Desktop\test.xlsx")